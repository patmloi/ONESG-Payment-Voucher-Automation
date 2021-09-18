from datetime import date
import dropbox
import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import os
import pandas as pd
from pandas import ExcelWriter
import typing

# EDIT
# Prepared by
prepared_by_full = "Preparer Name"
prepared_by = prepared_by_full.split()[0] # Preparer's first name

# Approved by
approved_by_full = "Approver Name"
approved_by = " ".join(approved_by_full.split()[1:]) # Approver's first name

# Dropbox
dbx_access_token = "Dropbox API Access Token"

# Email
# Sender email account
email_sender = "sender_email@gmail.com"
email_sender_pw = "sender_email_password"

# Receiver emails
email_receiver_main = "receiver_email@gmail.com"
email_receiver_cc = "receiver2@gmail.com, receiver3@gmail.com"

# Directory
directory = r"C:\Users\User1\Documents" # Location of root folder

# DO NOT EDIT
root_path = os.path.join(directory, "ONE Singapore")
fr_path = os.path.join(root_path, "Food Reimbursements")

class EFRefNos(): 
    def __init__(self, orders_no: int) -> None: 
        self.nos = [] 
        self.ref_nos = []
        today = date.today()
        year = today.strftime("%Y")

        for i in range(orders_no):
            no = input("EF Ref No. (Excluding year): ")
            self.nos.append(no)

            ref = f"EF {year}-{no}"
            self.ref_nos.append(ref)

        self.ref_nos_str = ", ".join(self.ref_nos)

        return 

    pass


class OSRefNo(): 
    def __init__(self, orders_no: int, ef: object, dbx_access_token: str) -> None:
        today = date.today()
        year = today.strftime("%y")

        # OS No 
        if orders_no == 1: # Single
            ef_nos = getattr(ef, "nos")
            self.no = ef_nos[0]

        
        else: # Multiple
            received_date = input("Date order was received (DD/MM, No spacing): ")
            received_date, received_month = received_date.split("/")

            # One multiple order in the received day (Default)
            self.no = f"{received_month}0{received_date}"

            total_multi_no = int(input("No. of multiple orders on recevied date: "))
            # Multiple multiple orders in the received day
            if total_multi_no > 1:
                lower_alpha = "abcdefghijklmnopqrstuvwxyz"
                multi_position_no = int(input("No. of multiple orders processed before current multiple order: "))
                add_letter = lower_alpha[multi_position_no]
                self.no = f"{self.no}{add_letter}"


        # OS Letter Code 
        if orders_no == 1: # Single  
            self.letter = "FR"
        else: # Multiple
            self.letter = "FMR"

        # OS Ref no
        self.ref_no = f"OS-{year}-{self.no}{self.letter}"

        # Repeated delivery for single case: Possible change in OS Ref no

        # Dropbox access
        dbx = dropbox.Dropbox(dbx_access_token)

        # Alphabet
        alpha_index = 0 
        alpha = "abcdefghijklmnopqrstuvwxyz"

        if orders_no == 1: 
            repeat = True 

            # Repeats until an unused reference no. is found.
            while repeat: 
                dbx_query = self.ref_no
                db_results = dbx.files_search_v2(dbx_query)
                dbx_matches = db_results.matches

                if dbx_matches == []:
                    repeat = False

                else: 
                    self.no = f"{self.no}{alpha[alpha_index]}"
                    self.ref_no = f"OS-{year}-{self.no}{self.letter}"
                    alpha_index += 1
                
        return 

    pass



class Volunteer(): 
    def __init__(self, root_path: str): 
        self.short = input("Volunteer's first name (name used when order was received): ")
        
        # Excel 
        self.root_path = root_path
        self.ref_file_path = os.path.join(root_path, "Volunteers.xlsx")
        self.wb = openpyxl.load_workbook(self.ref_file_path)
        self.ws = self.wb["Payment"]
        return

    def search(self): 
        '''
        Searches for a volunteer using their shorr name in the Search column. 
        If found, assigns attribute found as True, and assigns attribute row. 
        If not found, assigns attribute found as False.
        '''
        self.found = False

        for row in range(1, self.ws.max_row + 1): 
            if self.ws[row][0].value == self.short: 
                self.found = True
                self.row = row
                return 
            
        return 

    def write_details(self): 
        '''Creates a new workbook with details of existing and new volunteers.'''

        self.write_wb = openpyxl.Workbook()
        self.write_ws = self.write_wb.active
        self.write_ws.title = "Payment"


        for r in range(1, self.ws.max_row + 1): 
            for c in range(5):
                self.write_ws.cell(r, c+1, self.ws[r][c].value)


        self.write_ws.cell(self.row, 1, self.short)
        self.write_ws.cell(self.row, 2, self.full)
        self.write_ws.cell(self.row, 3, self.payment)
        self.write_ws.cell(self.row, 4, self.type_int)
        self.write_ws.cell(self.row, 5, self.details)

        self.write_wb.save(self.ref_file_path)
        self.write_wb.close()
        return 

    def assign(self): 
        '''
        Assigns full name, payment name, transaction type (boolean, full description) and payment details attributes. 
        If the volunteer was not found in the worksheet, writes short name and the above attributes to the worksheet.
        '''

        if self.found:
            self.full = self.ws[self.row][1].value
            self.payment = self.ws[self.row][2].value
            self.type_int = self.ws[self.row][3].value
            self.details = self.ws[self.row][4].value
            self.wb.close()

        else: 
            self.row = self.ws.max_row + 1 
            # Input details of new 
            self.full = input("Full name: ")
            self.payment = input("Payment name: ")
            self.type_int = int(input("Payment type (Bank Account: 0, PayNow: 1): "))
            self.details = input("Payment details: ")

            # Writes the details of the new volunteer on the worksheet.
            self.write_details()


        types = ["Transfer to Bank Account", "PayNow"]
        self.type = types[self.type_int]
        
        return 

    def volunteer(self): 
        '''
        Searches the worksheet and assigns Volunteer attributes:
        found, row, full, payment, type_bool, type, details, 
        and saves the Volunteer workbook before closing it.
        '''
        self.search()
        self.assign()

        return 
        
    pass



class Food(): 
    def __init__(self):
        self.retailer_abbrev = input("Retailer abbreviation (Giant: G, NTUC Fairprice: NTUC, Prime: P, Sheng Siong: SS): ")
        self.retailers = {
            "G": "Giant",
            "NTUC": "NTUC Fairprice", 
            "P": "Prime Supermarket", 
            "SS": "Sheng Siong"
        } # config.json

        self.retailer = self.retailers[self.retailer_abbrev]
        self.receipt_date = input("Receipt date (DD/MM/YYYY): ")
        self.receipt_no = input("Receipt no.: ")
        return 

    pass



class FreshFood(Food): 
    def __init__(self):
        super().__init__()
        self.desc = f"{self.retailer}: Fresh Foods"
        self.quantity = 1
        self.unit_price = float(input("Receipt amount (DD.CC): "))
        self.amount = self.unit_price
        return 

    pass



class Voucher(Food):
    def __init__(self):
        super().__init__()
        self.desc = f"{self.retailer}: Vouchers"
        self.v_desc = []
        self.type_no = int(input("No. of types of vouchers: "))
        return 

    def voucher_desc(self):
        value = int(input("Voucher value (No $): "))
        qty = int(input("No. of vouchers of the same value: "))
        total = qty * value

        first_voucher = input("First voucher no. : ")
        first_voucher_no = int(first_voucher)
        len_diff = len(first_voucher) - len(str(first_voucher_no))
        last_voucher_no = first_voucher_no + qty - 1
        extra_digits = len_diff * "0"
        voucher_no_range = f"{first_voucher} - {extra_digits}{last_voucher_no}"

        working = f"S${value} x {qty} = S${total}"

        voucher_desc = f"Voucher no. {voucher_no_range} ({working})"
        self.v_desc.append(voucher_desc)

        return qty, value, total

    def voucher(self):
        self.amount = 0 

        for i in range(self.type_no): 
            v = self.voucher_desc()
            self.amount += v[2]

        if self.type_no == 1: 
            self.quantity, self.unit_price = v[0], v[1]
        else: 
            self.quantity, self.unit_price = 1, self.amount
        return 
    
    pass



class FoodItems():
    def __init__(self): 
        self.fresh_no = int(input("No. of fresh food receipts: "))
        self.voucher_no = int(input("No. of voucher receipts: "))
        self.total_no = self.fresh_no + self.voucher_no 
        self.total_amount = 0
        self.list = []

        for f in range(self.fresh_no):
            fresh = FreshFood()
            self.total_amount += fresh.amount
            self.list.append(("F", fresh))
        
        for v in range(self.voucher_no):
            vouch = Voucher()
            vouch.voucher()
            self.total_amount += vouch.amount
            self.list.append(("V", vouch))


        return 
    
    pass



class Remark():
    def __init__(self, orders_no: int): 
        if orders_no == 1:
            indiv = input("Reference not uplaoded / Beneficiary pplied for themself (Y/N): ")

            if indiv == "Y": 
                self.remark = "Pending reference" # config.json
            else: 
                name = input("Name of referrer: ") 
                position = input("Position of referrer: ")
                org = input("Organisation of referrer: ")
                self.remark = f"Referred by: {name}, {position} from {org}"

        else: 
            self.remark = "Multiple EFs" # config.json

        return 

    pass



class Folder(): 
    def __init__(self, root_path: str, fr_path: str):
        self.fr_path = fr_path
        self.root_path = root_path
        return 

    def create(self, payment: object): 
        path = os.path.join(self.fr_path, payment.os.ref_no)
        if not os.path.isdir(path):
            os.makedirs(path)
            
        self.folder = path 
        return 

    pass 



class FR(): 
    def __init__(self, dbx_access_token: str, root_path: str, fr_path: str, prepared_by_full: str, prepared_by: str, approved_by_full: str, approved_by: str):
        # Constants 
        self.prepared_by_full = prepared_by_full
        self.prepared_by = prepared_by

        self.approved_by_full = approved_by_full
        self.approved_by = approved_by 

        # Date
        self.date = date.today() 
        self.doc_date = self.date.strftime("%d/%m/%Y")
        self.orders_no = int(input("No. of deliveries: "))
        
        # Objects
        self.ef = EFRefNos(self.orders_no)
        self.os = OSRefNo(self.orders_no, self.ef, dbx_access_token)

        if self.orders_no == 1: 
            self.beneficiary = input("Beneficiary's full name: ")
        else: 
            self.beneficiary = ""
            
        self.vol = Volunteer(root_path)
        self.vol.volunteer()
        self.items = FoodItems()
        self.re = Remark(self.orders_no)

        # Folder
        self.folder = Folder(root_path, fr_path)
        self.folder.create(self)

        return 

    pass

"""###Email"""

class Email(): 
    def __init__(self, fr: object, email_sender: str, email_sender_pw: str, email_receiver_main: str, email_receiver_cc: str, fr_path: str): 
        self.fr = fr
        self.os = self.fr.os.ref_no
        self.sender = email_sender
        self.sender_pw = email_sender_pw
        self.receiver_main = email_receiver_main
        self.receiver_cc = email_receiver_cc
        self.subject = f"Reimbursement for Fresh Food Purchase {self.os} for {self.fr.vol.full}"

        self.greeting = f"Hi {self.fr.approved_by},"
        self.text = f"Attached are the PV and supporting files for {self.os}."
        self.sign = f"Warm regards,\n{self.fr.prepared_by}"

        self.fr_path = fr_path

    def create(self):
        '''First creates the email, with:
        sender, receivers (main and cc), subject and body text.'''

        message= MIMEMultipart()
        message["From"] = self.sender
        message["To"] = self.receiver_main
        message["Cc"] = self.receiver_cc
        message["Subject"] = self.subject
        
        # Additional message 
        text_add_bool = input("Additional message? (Y/N) ")
        if text_add_bool.lower() == "y": 
            text_add = input("Additional message: ")
            self.body = f"{self.greeting}\n\n{self.text}\n\n{text_add}\n\n{self.sign}"
        else: 
            self.body = f"{self.greeting}\n\n{self.text}\n\n{self.sign}"

        message.attach(MIMEText(self.body, "plain"))
        self.email = message 

        return

    def attachment_paths(self):
        '''Creates a list of files to be attached to the email, 
        each file represented as a tuple (file name, file path).'''

        order_path = self.os
        full_path = os.path.join(self.fr_path, order_path)

        all_files = os.listdir(full_path)

        attach_files = []
        for f in all_files: 
            if f[0] in ["O", "P"]:
                file_path = os.path.join(full_path, f)
                attach_files.append((f, file_path))

        self.files = attach_files

        attach_files = []
        for f in all_files: 
            if f[0] in ["O", "P"]:
                file_path = os.path.join(full_path, f)
                attach_files.append((f, file_path))

        self.files = attach_files
        return 

    def attach(self, file_name, file_path):
        '''Attaches a file to the email.'''

        attachment = open(file_path, "rb")
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {file_name}",
        )

        self.email.attach(part)
        return 

    def attachments(self): 
        '''Attaches all files to the email.'''

        self.attachment_paths()
        for a in self.files: 
            file_name, file_path = a
            self.attach(file_name, file_path)
        return 

    def send(self): 
        '''Logs into the sender email account and sends the email.'''
        # Login
        context = ssl.create_default_context()
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.ehlo()  # Can be omitted
        server.starttls(context=context)
        server.ehlo()  # Can be omitted
        server.login(self.sender, self.sender_pw)

        # Send mail
        server.send_message(self.email)

        return 

    def email(self): 
        '''Creates an email with attachments, and sends it.'''
        self.create()
        self.attachments()
        self.send()
        return

"""###Payment Voucher"""

class PV():
    def __init__(self, payment: object): 
        self.payment = payment
        self.purpose = "Reimbursement -  ONE EF Fresh Food Programme" # Determined with os.letter (config.json)
        return 

    def template(self): 
        '''Creates a new order PV from the template PV.''' 

        # Template file paths
        temp_file_path = "Payment Voucher - Template.xlsx"
        temp_full_path = os.path.join(self.payment.folder.root_path, temp_file_path)
        # New file paths 
        new_file_path = f"Payment Voucher - {self.payment.os.ref_no} - {self.payment.vol.full}.xlsx"
        new_full_path = os.path.join(self.payment.folder.folder, new_file_path)
        self.full_path = new_full_path

        # Copying from template
        src = open(temp_full_path, 'rb')
        dst = open(new_full_path, 'wb').write(src.read())

        return 

    def __write_os(self): 
        self.ws.cell(row=4, column=7).value = self.payment.os.ref_no
        self.ws.cell(row=5, column=7).value = self.payment.doc_date
        return
    
    def __write_payee(self): 
        self.ws.cell(row=6, column=2).value = self.payment.vol.payment
        self.ws.cell(row=7, column=2).value = self.payment.beneficiary 
        self.ws.cell(row=8, column=2).value = self.payment.vol.type

        if self.payment.vol.type_int == 0: 
            self.ws.cell(row=9, column=2).value = self.payment.vol.details
        else: 
            self.ws.cell(row=10, column=2).value = self.payment.vol.details

        return 

    def __write_purpose(self): 
        self.ws.cell(row=12, column=2).value = self.purpose
        self.ws.cell(row=13, column=2).value = self.payment.ef.ref_nos_str
        return 

    def __write_items(self): 
        item_no = 0
        row_no = 15

        for i in range(self.payment.items.total_no): 
            item_no += 1
            row_no += 1
            
            item_code, item_obj = self.payment.items.list[i]
            
            # Item row 
            row_content = [
                item_no,
                item_obj.desc,
                item_obj.receipt_date,
                item_obj.receipt_no, 
                item_obj.quantity, 
                item_obj.unit_price,
                item_obj.amount

            ]

            for c in range(len(row_content)):
                self.ws.cell(row=row_no, column = c+1).value = row_content[c]

            # Voucher description(s)
            if item_code == "V": 
                for v in range(len(item_obj.v_desc)):
                    row_no += 1
                    self.ws.cell(row=row_no, column = 2).value = item_obj.v_desc[v]

        self.ws.cell(row=28, column = 7).value = self.payment.items.total_amount
        return 

    def __write_remark(self): 
        remark = self.payment.re.remark

        row_height = 12.75
        line_chara_count = 45

        line_no = len(remark) / 45
        if line_no < 1: 
            line_no = 1
            
        height = row_height * line_no


        self.ws.cell(row=30, column=2).value = remark
        rd = self.ws.row_dimensions[25]
        rd.height = height

        return 

    def __write_declare(self): 
        prep = self.payment.prepared_by_full
        app = self.payment.approved_by_full 
        # Approved by (Declaration)
        self.ws.cell(row=35, column=5).value = app

        # Prepared by (Row)
        self.ws.cell(row=38, column=2).value = prep
        self.ws.cell(row=38, column=5).value = self.payment.doc_date
        # Approved by (Row)
        self.ws.cell(row=39, column=2).value = app
        return 

    def write(self):
        self.wb = openpyxl.load_workbook(self.full_path)
        self.ws = self.wb.active
        
        self.__write_os()
        self.__write_payee()
        self.__write_purpose()
        self.__write_items()
        self.__write_remark()
        self.__write_declare()

        self.wb.save(self.full_path)
        self.wb.close
        return 

    def pv(self): 
        self.template()
        self.write()
        return 

    pass

"""### Dropbox"""

class DBX(): 
    def __init__(self, dbx_access_token: str, fr_path: str): 
        self.access_token = dbx_access_token
        self.dbx = dropbox.Dropbox(self.access_token)

        self.fr_path = fr_path
        return 

    def create_folder(self):
        self.dropbox_folder = f"/ONE (SINGAPORE)/Campaigns/2021/ONE Emergency Fund/Food Receipts 11 Apr - 31 Dec 2021 (NEW format)/{self.payment.os.ref_no} - {self.payment.vol.full}"
        self.dbx.files_create_folder(self.dropbox_folder)
        return

    def upload(self, payment: object): 
        '''Uploads all files in the local folder 
        into the newly created Dropbox folder,
        except for the Payment Voucher as 
        it is still pending confirmation.''' 
        
        # Creating new Dropbox folder 
        self.payment = payment
        self.local_folder = self.payment.folder.folder
        self.create_folder()
        
        # Uploading files into Dropbox 
        files = os.listdir(self.local_folder)
        for f in range(len(files)):
            file_name = files[f]

            if file_name[0] != "P": 
                local_file_path = os.path.join(self.local_folder, file_name)
                dbx_file_path = f"{self.dropbox_folder}/{file_name}"
                file = open(local_file_path, "rb")
                self.dbx.files_upload(file.read(), dbx_file_path)
        return 

    def search_folder(self): 
        '''Returns the path of the local file being searched for.'''
        today = date.today()
        year = today.strftime("%y")
        partial_ref_no = input("OS Ref. no., without OS-20XX: ")
        ref_no = f"OS-{year}-{partial_ref_no}"
        folder = os.path.join(self.fr_path, ref_no)
        return folder

    def search_pv(self, folder: str): 
        '''Returns the name and path of the Payment Voucher being searched for.'''
        files = os.listdir(folder)
        for f in files: 
            if f[0] == "P":
                name = f
                path = os.path.join(folder, f)
        return name, path

    def dbx_folder_path(self, name):
        '''Returns the path of a Dropbox folder.'''
        parent = "/ONE (SINGAPORE)/Campaigns/2021/ONE Emergency Fund/Food Receipts 11 Apr - 31 Dec 2021 (NEW format)/"
        child = name[18:-5]
        dbx_folder = f"{parent}{child}"
        return dbx_folder

    def dbx_file_path(self, folder, file): 
        '''Returns the path of a Dropbox file.'''
        return f"{folder}/{file}"


    def upload_pv(self):
        '''Locates the Dropbox folder and uploads the approved PV there.'''
        local_folder_path = self.search_folder()
        name, path = self.search_pv(local_folder_path)
        dbx_folder = self.dbx_folder_path(name)
        dbx_file =  self.dbx_file_path(dbx_folder, name)

        file = open(path, "rb")
        self.dbx.files_upload(file.read(), dbx_file)

        return 
    
    def create_folder_existing(self, folder:str):
        ''' Creates a new Dropbox folder for a manually created local folder.'''
        name = self.search_pv(folder)[0]
        dbx_folder = self.dbx_folder_path(name)
        self.dbx.files_create_folder_v2(dbx_folder)
        return dbx_folder
        
    def upload_existing(self): 
        '''Uploads the contents of a manually created local folder into Dropbox.'''
        local_folder = self.search_folder()
        dbx_folder = self.create_folder_existing(local_folder)
        files = os.listdir(local_folder)

        for f in range(len(files)): 
            dbx_file_path = self.dbx_file_path(dbx_folder, files[f])
            local_file_path = os.path.join(local_folder, files[f])
            local_file_open = open(local_file_path, "rb")
            self.dbx.files_upload(local_file_open.read(), dbx_file_path, autorename = True)
            
        return 

    pass


fr = FR(dbx_access_token, root_path, fr_path, prepared_by_full, prepared_by, approved_by_full, approved_by)
pv = PV(fr)
pv.pv()

downloaded_files_check = input("Downloaded all files? (Y/N) ")

while downloaded_files_check.lower == "n": 
    downloaded_files_check = input("Downloaded all files? (Y/N) ")

if downloaded_files_check.lower == "y": 
    e = Email(fr, email_sender, email_sender_pw, email_receiver_main, email_receiver_cc, fr_path)  
    e.email()

    dbx = DBX(dbx_access_token, fr_path)
    dbx.upload(fr)